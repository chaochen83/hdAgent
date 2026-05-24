from __future__ import annotations

import hashlib
import os
import re
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..core.config import settings
from .chunking import chunk_sheet_rows, chunk_text, estimate_token_count


@dataclass
class ParsedKnowledge:
    title: str
    knowledge_type: str
    source_type: str
    source_name: str
    raw_text: str
    chunks: list[dict[str, Any]]
    metadata: dict[str, Any]
    file_ext: str | None = None
    mime_type: str | None = None
    checksum_sha256: str | None = None
    file_size: int | None = None
    storage_path: str | None = None
    file_name: str | None = None


_STORAGE_DIR = Path(settings.knowledge_storage_dir).expanduser()
_CAD_EXTENSIONS = {".sch", ".pro", ".brd", ".kicad_sch", ".kicad_pro", ".kicad_pcb"}
_MCU_KEYWORDS = ("ESP32", "ESP8266", "STM32", "RP2040", "NRF", "ATSAMD", "ATMEGA", "CH32", "GD32")
_PASSIVE_PREFIXES = ("R", "C", "L", "FB")


def _ensure_storage_dir() -> Path:
    _STORAGE_DIR.mkdir(parents=True, exist_ok=True)
    return _STORAGE_DIR


def _sha256(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _decode_text(data: bytes) -> str:
    for encoding in ("utf-8", "utf-8-sig", "gbk", "gb18030", "latin-1"):
        try:
            return data.decode(encoding)
        except Exception:
            continue
    return data.decode("utf-8", errors="ignore")


def _render_workbook(data: bytes) -> tuple[str, list[dict[str, Any]], dict[str, Any]]:
    workbook = load_workbook(filename=BytesIO(data), data_only=True, read_only=True)
    all_texts: list[str] = []
    all_chunks: list[dict[str, Any]] = []
    sheet_names: list[str] = []
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        sheet_names.append(sheet_name)
        rows: list[list[str]] = []
        for row in ws.iter_rows(values_only=True):
            values = ["" if value is None else str(value) for value in row]
            if any(value.strip() for value in values):
                rows.append(values)
        if not rows:
            continue
        rendered_lines = [" | ".join([cell.strip() for cell in row if cell and cell.strip()]) for row in rows]
        all_texts.append(f"Sheet: {sheet_name}\n" + "\n".join(rendered_lines))
        all_chunks.extend(chunk_sheet_rows(sheet_name, rows, chunk_size=settings.knowledge_chunk_size))
    return "\n\n".join(all_texts).strip(), all_chunks, {"sheet_names": sheet_names}


def _normalize_pdf_text(text: str) -> str:
    normalized = (text or "").replace("\r\n", "\n").replace("\r", "\n")
    normalized = re.sub(r"[ \t]+", " ", normalized)
    normalized = re.sub(r"\n{3,}", "\n\n", normalized)
    return normalized.strip()


def _extract_parameter_candidates(text: str, *, limit: int = 40) -> list[dict[str, str]]:
    patterns = [
        re.compile(r"^\s*([A-Za-z][A-Za-z0-9 /_+().,-]{1,80}?)\s*[:：]\s*([^\n]{1,120})\s*$"),
        re.compile(r"^\s*([A-Za-z][A-Za-z0-9 /_+().,-]{1,80}?)\s{2,}([^\n]{1,120})\s*$"),
        re.compile(r"^\s*([A-Za-z][A-Za-z0-9 /_+().,-]{1,80}?)\s*[.·•]{2,}\s*([^\n]{1,120})\s*$"),
        re.compile(r"^\s*[•*-]\s*([A-Za-z][A-Za-z0-9 /_+().,-]{1,80}?)\s+of\s+([^\n]{1,120})\s*$"),
        re.compile(r"^\s*[•*-]\s*([A-Za-z][A-Za-z0-9 /_+().,-]{1,80}?)\s+from\s+([^\n]{1,120})\s*$"),
    ]
    candidates: list[dict[str, str]] = []
    seen: set[tuple[str, str]] = set()

    for raw_line in text.splitlines():
        line = raw_line.strip()
        if len(line) < 4 or len(line) > 160:
            continue
        if not any(ch.isdigit() for ch in line):
            continue
        match = None
        for pattern in patterns:
            match = pattern.match(line)
            if match:
                break
        if not match:
            continue
        name = re.sub(r"\s+", " ", match.group(1)).strip(" :-\t")
        value = re.sub(r"\s+", " ", match.group(2)).strip()
        if len(name) < 2 or len(value) < 1:
            continue
        if len(name.split()) > 8:
            continue
        lowered_name = name.lower()
        if lowered_name in {"page", "table", "figure", "revision"}:
            continue
        if lowered_name.startswith(("table ", "figure ", "note ", "page ")):
            continue
        if lowered_name.startswith(("the ", "this ", "these ", "for ")):
            continue
        if "...." in line or "..." in line:
            continue
        dedupe_key = (name.lower(), value.lower())
        if dedupe_key in seen:
            continue
        seen.add(dedupe_key)
        candidates.append({"name": name, "value": value})
        if len(candidates) >= limit:
            break

    return candidates


def _render_pdf(data: bytes, *, file_name: str) -> tuple[str, list[dict[str, Any]], dict[str, Any]]:
    try:
        from pypdf import PdfReader
    except ImportError as exc:
        raise RuntimeError("PDF parsing requires `pypdf`. Please install project dependencies again.") from exc

    reader = PdfReader(BytesIO(data))
    page_texts: list[str] = []
    for page_index, page in enumerate(reader.pages, start=1):
        text = _normalize_pdf_text(page.extract_text() or "")
        if not text:
            continue
        page_texts.append(f"[Page {page_index}]\n{text}")

    raw_text = "\n\n".join(page_texts).strip()
    spec_candidates = _extract_parameter_candidates(raw_text)
    chunks = chunk_text(
        raw_text,
        chunk_size=settings.knowledge_chunk_size,
        chunk_overlap=settings.knowledge_chunk_overlap,
        prefix=f"Source PDF: {file_name}",
        base_metadata={"content_type": "pdf", "page_count": len(reader.pages)},
    )
    return raw_text, chunks, {"page_count": len(reader.pages), "spec_candidates": spec_candidates}


def _looks_like_mcu_part(part: dict[str, str]) -> bool:
    part_name = str(part.get("name", "") or "").upper()
    if part_name and not part_name.startswith(("U", "IC", "MOD", "ESP")):
        return False
    haystack = " ".join(
        str(part.get(key, "") or "")
        for key in ("name", "library", "deviceset", "device", "value")
    ).upper()
    return any(keyword in haystack for keyword in _MCU_KEYWORDS)


def _looks_like_mcu_pin(pin_name: str) -> bool:
    token = (pin_name or "").upper()
    return bool(
        re.fullmatch(
            r"(?:IO|GPIO)\d+|TXD\d*|RXD\d*|SDA\d*|SCL\d*|MISO|MOSI|SCLK|BCLK|LRCLK|LRCK|WS|CLK|CS|RST|INT#?|A\d+|D\d+",
            token,
        )
    )


def _humanize_signal_name(name: str) -> str:
    normalized = re.sub(r"[_\-]+", " ", (name or "").strip()).strip()
    if not normalized:
        return "Unlabeled Signal"
    special_cases = {
        "SDA": "I2C SDA",
        "SCL": "I2C SCL",
        "RXD": "UART RXD",
        "TXD": "UART TXD",
        "HMISO": "SPI MISO",
        "HMOSI": "SPI MOSI",
        "SCLK": "SPI SCLK",
        "SPI NCS2": "SPI CS2",
        "LCD RS": "LCD RS/DC",
        "LCD RST": "LCD RST",
        "I2S LRCK": "I2S LRCK",
        "I2S SCLK": "I2S SCLK",
        "I2S DATA": "I2S DATA",
    }
    upper = normalized.upper()
    for key, value in special_cases.items():
        if upper == key:
            return value
    if upper.startswith("CSI "):
        return normalized.replace("CSI ", "Camera CSI ", 1)
    return normalized


def _is_noise_endpoint(part_name: str, part_meta: dict[str, str] | None) -> bool:
    if not part_name:
        return True
    if part_name.startswith(("+", "GND", "SUPPLY", "U$")):
        return True
    upper_name = part_name.upper()
    if upper_name.startswith(_PASSIVE_PREFIXES):
        return True
    deviceset = (part_meta or {}).get("deviceset", "").upper()
    if any(token in deviceset for token in ("GND", "CAP", "RESISTOR", "R*", "C*", "INDUCT", "DIODE", "LED", "TVS")):
        return True
    return False


def _render_eagle_schematic(data: bytes, *, file_name: str) -> tuple[str, list[dict[str, Any]], dict[str, Any]]:
    root = ET.parse(BytesIO(data)).getroot()
    parts = {part.attrib.get("name", ""): dict(part.attrib) for part in root.findall(".//parts/part")}
    mcu_parts = {name: meta for name, meta in parts.items() if _looks_like_mcu_part(meta)}

    hardware_resources: list[dict[str, Any]] = []
    net_summaries: list[str] = []
    for net in root.findall(".//nets/net"):
        net_name = (net.attrib.get("name") or "").strip()
        pinrefs = [dict(pinref.attrib) for pinref in net.findall(".//pinref")]
        pinref_labels = [f"{item.get('part')}.{item.get('pin')}" for item in pinrefs if item.get("part") and item.get("pin")]
        if net_name and pinref_labels:
            net_summaries.append(f"- {net_name}: {', '.join(pinref_labels)}")

        for pinref in pinrefs:
            part_name = pinref.get("part") or ""
            pin_name = pinref.get("pin") or ""
            if part_name not in mcu_parts or not _looks_like_mcu_pin(pin_name):
                continue

            signal_tokens = [token.strip() for token in re.split(r"[\\/]", net_name) if token.strip()]
            function_tokens = [token for token in signal_tokens if not re.fullmatch(r"(?:IO|GPIO)\d+", token.upper())]
            function_name = _humanize_signal_name("/".join(function_tokens) or net_name or pin_name)
            connected_refs = []
            for item in pinrefs:
                other_part = item.get("part") or ""
                other_pin = item.get("pin") or ""
                if other_part == part_name and other_pin == pin_name:
                    continue
                if _is_noise_endpoint(other_part, parts.get(other_part)):
                    continue
                connected_refs.append(f"{other_part}.{other_pin}")
            connected_refs = sorted(dict.fromkeys(connected_refs))
            hardware_resources.append(
                {
                    "function": function_name,
                    "gpio_pin": pin_name,
                    "mcu_part": part_name,
                    "net_name": net_name or pin_name,
                    "connected_refs": connected_refs,
                }
            )

    deduped_resources: list[dict[str, Any]] = []
    seen_keys: set[tuple[str, str, str]] = set()
    for item in hardware_resources:
        key = (item["mcu_part"], item["gpio_pin"], item["net_name"])
        if key in seen_keys:
            continue
        seen_keys.add(key)
        deduped_resources.append(item)

    deduped_resources.sort(
        key=lambda item: (
            item["mcu_part"],
            int(re.search(r"\d+", item["gpio_pin"]).group()) if re.search(r"\d+", item["gpio_pin"]) else 999,
            item["gpio_pin"],
        )
    )

    board_title = ""
    text_nodes = root.findall(".//plain/text")
    if text_nodes:
        board_title = " | ".join(text.text.strip() for text in text_nodes if text.text and "ESP32" in text.text.upper())[:200]

    resource_lines = [
        f"- {item['function']}: {item['mcu_part']}.{item['gpio_pin']} (net {item['net_name']})"
        + (f" -> {', '.join(item['connected_refs'])}" if item["connected_refs"] else "")
        for item in deduped_resources
    ]
    summary_lines = [
        f"CAD file: {file_name}",
        "CAD format: Eagle schematic (XML)",
        f"Board title: {board_title or 'Unknown'}",
        f"Detected MCU parts: {', '.join(f'{name} ({meta.get('deviceset') or meta.get('value') or 'MCU'})' for name, meta in mcu_parts.items()) or 'None'}",
        f"Part count: {len(parts)}",
        f"Net count: {len(root.findall('.//nets/net'))}",
        "",
        "Hardware Resource Mapping:",
        *resource_lines,
        "",
        "Net Connectivity Summary:",
        *net_summaries,
    ]
    raw_text = "\n".join(line for line in summary_lines if line is not None).strip()
    chunks = chunk_text(
        raw_text,
        chunk_size=settings.knowledge_chunk_size,
        chunk_overlap=settings.knowledge_chunk_overlap,
        prefix=f"Source CAD: {file_name}",
        base_metadata={"content_type": "cad", "cad_format": "eagle_schematic"},
    )
    return raw_text, chunks, {
        "content_type": "cad",
        "cad_format": "eagle_schematic",
        "part_count": len(parts),
        "net_count": len(root.findall('.//nets/net')),
        "board_title": board_title or None,
        "detected_mcu_parts": [
            {"part_name": name, "deviceset": meta.get("deviceset") or meta.get("value") or "MCU"}
            for name, meta in mcu_parts.items()
        ],
        "hardware_resources": deduped_resources,
    }


def _render_cad_file(data: bytes, *, file_name: str) -> tuple[str, list[dict[str, Any]], dict[str, Any]]:
    ext = os.path.splitext(file_name or "")[1].lower()
    if ext == ".sch":
        try:
            root = ET.parse(BytesIO(data)).getroot()
            if root.tag == "eagle":
                return _render_eagle_schematic(data, file_name=file_name)
        except ET.ParseError:
            pass

    raw_text = _decode_text(data).strip()
    chunks = chunk_text(
        raw_text,
        chunk_size=settings.knowledge_chunk_size,
        chunk_overlap=settings.knowledge_chunk_overlap,
        prefix=f"Source CAD: {file_name}",
        base_metadata={"content_type": "cad", "cad_format": "text"},
    )
    return raw_text, chunks, {"content_type": "cad", "cad_format": "text", "file_ext": ext.lstrip(".")}


def detect_knowledge_type_from_file_name(file_name: str) -> str:
    ext = os.path.splitext(file_name or "")[1].lower()
    if ext == ".txt":
        return "txt"
    if ext == ".pdf":
        return "pdf"
    if ext in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return "excel"
    if ext in _CAD_EXTENSIONS:
        return "cad"
    raise ValueError("Only .txt, .pdf, .xlsx and common CAD design files (.sch/.pro/.brd/.kicad_*) are supported for knowledge ingestion right now.")


def store_upload_file(*, file_name: str, data: bytes, mime_type: str | None = None) -> dict[str, Any]:
    ext = os.path.splitext(file_name or "")[1].lower()
    knowledge_type = detect_knowledge_type_from_file_name(file_name)
    checksum = _sha256(data)
    storage_dir = _ensure_storage_dir()
    storage_path = storage_dir / f"{checksum}{ext}"
    storage_path.write_bytes(data)
    return {
        "knowledge_type": knowledge_type,
        "file_ext": ext.lstrip("."),
        "mime_type": mime_type,
        "checksum_sha256": checksum,
        "file_size": len(data),
        "storage_path": str(storage_path),
    }


def read_stored_upload(storage_path: str) -> bytes:
    path = Path(storage_path)
    if not path.is_file():
        raise FileNotFoundError(f"Stored upload not found: {storage_path}")
    return path.read_bytes()


def parse_text_input(*, title: str, text: str) -> ParsedKnowledge:
    raw_text = (text or "").strip()
    chunks = chunk_text(
        raw_text,
        chunk_size=settings.knowledge_chunk_size,
        chunk_overlap=settings.knowledge_chunk_overlap,
    )
    return ParsedKnowledge(
        title=title.strip() or "未命名文本知识",
        knowledge_type="text",
        source_type="text",
        source_name=title.strip() or "text",
        raw_text=raw_text,
        chunks=chunks,
        metadata={"input_method": "text"},
    )


def parse_upload(*, file_name: str, data: bytes, mime_type: str | None = None) -> ParsedKnowledge:
    ext = os.path.splitext(file_name or "")[1].lower()

    if ext == ".txt":
        raw_text = _decode_text(data).strip()
        chunks = chunk_text(
            raw_text,
            chunk_size=settings.knowledge_chunk_size,
            chunk_overlap=settings.knowledge_chunk_overlap,
        )
        knowledge_type = "txt"
        metadata = {"content_type": "text"}
    elif ext == ".pdf":
        raw_text, chunks, metadata = _render_pdf(data, file_name=file_name)
        knowledge_type = "pdf"
    elif ext in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        raw_text, chunks, metadata = _render_workbook(data)
        knowledge_type = "excel"
    elif ext in _CAD_EXTENSIONS:
        raw_text, chunks, metadata = _render_cad_file(data, file_name=file_name)
        knowledge_type = "cad"
    else:
        raise ValueError("Only .txt, .pdf, .xlsx and common CAD design files (.sch/.pro/.brd/.kicad_*) are supported for knowledge ingestion right now.")

    stored = store_upload_file(file_name=file_name, data=data, mime_type=mime_type)
    title = os.path.splitext(os.path.basename(file_name or "知识文件"))[0] or "知识文件"
    return ParsedKnowledge(
        title=title,
        knowledge_type=knowledge_type,
        source_type="file",
        source_name=file_name,
        raw_text=raw_text,
        chunks=chunks,
        metadata=metadata,
        file_ext=stored["file_ext"],
        mime_type=stored["mime_type"],
        checksum_sha256=stored["checksum_sha256"],
        file_size=stored["file_size"],
        storage_path=stored["storage_path"],
        file_name=file_name,
    )


def summarize_parsed_knowledge(parsed: ParsedKnowledge) -> dict[str, Any]:
    return {
        "chunk_count": len(parsed.chunks),
        "token_count": estimate_token_count(parsed.raw_text),
        "metadata": parsed.metadata,
    }
